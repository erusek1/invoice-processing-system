#!/usr/bin/env python3
"""
Excel Manager Module
------------------
Handles writing invoice data to Excel workbooks organized by date.
"""

import os
import pandas as pd
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

class ExcelManager:
    def __init__(self, output_file="invoice_data.xlsx"):
        """
        Initialize Excel manager.
        
        Args:
            output_file: Path to the Excel file for output
        """
        self.output_file = output_file
        self.workbook = self._load_or_create_workbook()
        self.data_by_day = {}
        
    def _load_or_create_workbook(self):
        """
        Load existing workbook or create a new one.
        
        Returns:
            openpyxl Workbook object
        """
        if os.path.exists(self.output_file):
            try:
                return openpyxl.load_workbook(self.output_file)
            except Exception as e:
                print(f"Error loading workbook: {str(e)}")
                print("Creating a new workbook instead.")
                return openpyxl.Workbook()
        else:
            return openpyxl.Workbook()
    
    def add_invoice(self, invoice_data):
        """
        Add an invoice to the appropriate daily sheet.
        
        Args:
            invoice_data: Dictionary containing invoice information
        """
        # Use the invoice date if available, otherwise use today's date
        date_key = invoice_data.get('date')
        if not date_key:
            date_key = datetime.now().strftime('%Y-%m-%d')
            
        # Format sheet name as MM-DD-YYYY
        try:
            date_obj = datetime.strptime(date_key, '%Y-%m-%d')
            sheet_name = date_obj.strftime('%m-%d-%Y')
        except ValueError:
            # Fall back to today's date if parsing fails
            sheet_name = datetime.now().strftime('%m-%d-%Y')
            
        # Add to our data collection
        if sheet_name not in self.data_by_day:
            self.data_by_day[sheet_name] = []
            
        self.data_by_day[sheet_name].append(invoice_data)
        
        # Update Excel immediately
        self._update_sheet(sheet_name, self.data_by_day[sheet_name])
    
    def _update_sheet(self, sheet_name, invoices):
        """
        Update or create a worksheet for a specific date.
        
        Args:
            sheet_name: Name of the sheet (date formatted as MM-DD-YYYY)
            invoices: List of invoice dictionaries
        """
        # Create sheet if it doesn't exist
        if sheet_name not in self.workbook.sheetnames:
            sheet = self.workbook.create_sheet(title=sheet_name)
            self._setup_sheet_headers(sheet)
        else:
            sheet = self.workbook[sheet_name]
            
        # Clear existing data (except headers)
        for row in range(2, sheet.max_row + 1):
            for col in range(1, sheet.max_column + 1):
                sheet.cell(row=row, column=col).value = None
                
        # Add invoice data
        for i, invoice in enumerate(invoices, start=2):  # Start at row 2 (after headers)
            sheet.cell(row=i, column=1).value = invoice.get('invoice_number', '')
            sheet.cell(row=i, column=2).value = invoice.get('date', '')
            sheet.cell(row=i, column=3).value = invoice.get('job_name', '')
            sheet.cell(row=i, column=4).value = invoice.get('supply_house', '')
            sheet.cell(row=i, column=5).value = invoice.get('total_cost', 0)
            sheet.cell(row=i, column=6).value = invoice.get('job_cost', 0)
            
            # Format currency cells
            sheet.cell(row=i, column=5).number_format = '$#,##0.00'
            sheet.cell(row=i, column=6).number_format = '$#,##0.00'
            
        # Add totals row
        total_row = len(invoices) + 2
        sheet.cell(row=total_row, column=4).value = 'TOTALS:'
        sheet.cell(row=total_row, column=5).value = f'=SUM(E2:E{total_row-1})'
        sheet.cell(row=total_row, column=6).value = f'=SUM(F2:F{total_row-1})'
        
        # Format totals row
        for col in range(4, 7):
            cell = sheet.cell(row=total_row, column=col)
            cell.font = Font(bold=True)
            if col >= 5:
                cell.number_format = '$#,##0.00'
                
        # Auto-adjust column widths
        self._adjust_column_widths(sheet)
    
    def _setup_sheet_headers(self, sheet):
        """
        Set up headers and formatting for a new sheet.
        
        Args:
            sheet: openpyxl worksheet object
        """
        headers = [
            'Invoice #', 
            'Date', 
            'Job Name', 
            'Supplier', 
            'Total Cost', 
            'Job Cost'
        ]
        
        # Add headers
        for i, header in enumerate(headers, start=1):
            cell = sheet.cell(row=1, column=i)
            cell.value = header
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color='DDDDDD', end_color='DDDDDD', fill_type='solid')
            cell.alignment = Alignment(horizontal='center')
            
        # Format columns
        sheet.column_dimensions['A'].width = 12  # Invoice #
        sheet.column_dimensions['B'].width = 12  # Date
        sheet.column_dimensions['C'].width = 30  # Job Name
        sheet.column_dimensions['D'].width = 20  # Supplier
        sheet.column_dimensions['E'].width = 12  # Total Cost
        sheet.column_dimensions['F'].width = 12  # Job Cost
    
    def _adjust_column_widths(self, sheet):
        """
        Auto-adjust column widths based on content.
        
        Args:
            sheet: openpyxl worksheet object
        """
        for col in range(1, sheet.max_column + 1):
            max_length = 0
            column = get_column_letter(col)
            
            for row in range(1, sheet.max_row + 1):
                cell = sheet.cell(row=row, column=col)
                if cell.value:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                        
            adjusted_width = (max_length + 2) if max_length > 0 else 10
            sheet.column_dimensions[column].width = min(adjusted_width, 50)  # Cap at 50
    
    def save(self):
        """Save the workbook to disk."""
        try:
            # Create directory if needed
            output_dir = os.path.dirname(self.output_file)
            if output_dir and not os.path.exists(output_dir):
                os.makedirs(output_dir)
                
            self.workbook.save(self.output_file)
            return True
        except Exception as e:
            print(f"Error saving workbook: {str(e)}")
            return False